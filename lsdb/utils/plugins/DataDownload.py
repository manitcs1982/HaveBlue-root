class DataDownload():
    import json
    import zipfile
    import numpy as np
    from PIL import Image, ImageOps
    import datetime
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO


    from django.db import IntegrityError, transaction
    from django.http import FileResponse, HttpResponse
    from django.utils import timezone
    from lsdb.utils import ExcelFile
    from django.db.models import Q, Sum, Max, Min, F

    from rest_framework.response import Response

    from lsdb.serializers.WorkOrderSerializer import WorkOrderSerializer

    from lsdb.models import AzureFile
    from lsdb.models import Disposition
    from lsdb.models import MeasurementResult
    from lsdb.models import ProcedureResult
    from lsdb.models import TestSequenceDefinition
    from lsdb.models import Project
    from lsdb.models import Unit
    from lsdb.models import WorkOrder

    from lsdb.utils.ObjFromUrl import obj_from_url

    def __init__(self):
        self.ids = [1]

    def _test(self):
        pass

    def _run(self, workorder_ids=[], unit_ids=[], procedure_ids=[], TSD_ids=[]):
        if workorder_ids:
            queryset = self.WorkOrder.objects.filter(id__in=workorder_ids)
            mem_zip = self.BytesIO()

            with self.zipfile.ZipFile(mem_zip, mode='w', compression=self.zipfile.ZIP_DEFLATED) as zf:
                for work_order in queryset:

                    sheets = []

                    excel_file_units = self.ExcelFile()
                    unit_sheet = excel_file_units.workbook.active
                    unit_sheet.append(["Manufacturer","Technology","Model","Serial Number","Calibration Device Used","Maximum Voltage[V]","Width [mm]","Height [mm]","Pmax [W]","VOC [V]","VMP[V]","ISC [A]","IMP [A]"])
                    sheets.append(unit_sheet)

                    excel_file_vi = self.ExcelFile()
                    vi_sheet = excel_file_vi.workbook.active
                    vi_sheet.append(["Serial Number", "TSD", "LEG", "Defect", "Category", "Notes", "Images:"])
                    sheets.append(vi_sheet)

                    excel_file_el = self.ExcelFile()
                    el_sheet = excel_file_el.workbook.active
                    el_sheet.append(["Serial Number", "TSD", "LEG", "Aperture", "ISO", "Exposure Count", "Injection Current", "Exposure Time"])
                    sheets.append(el_sheet)

                    excel_file_wl = self.ExcelFile()
                    wl_sheet = excel_file_wl.workbook.active
                    wl_sheet.append(["Serial Number", "TSD", "LEG", "Insulation Resistance", "Passed?", "Test Voltage", "Leakage Current", "Current Trip Setpoint", "Water Temperature"])
                    sheets.append(wl_sheet)

                    excel_file_dt = self.ExcelFile()
                    dt_sheet = excel_file_dt.workbook.active
                    dt_sheet.append(["Serial Number", "TSD", "LEG", "Forward Voltage", "Reverse Voltage", "Pass?"])
                    sheets.append(dt_sheet)

                    excel_file_flash = self.ExcelFile()
                    flash_sheet = excel_file_flash.workbook.active
                    flash_sheet.append(["Serial Number", "TSD", "LEG", "Pmp", "Voc", "Vmp", "Isc", "Imp", "Irradiance", "Temperature"])
                    sheets.append(flash_sheet)

                    excel_file_color = self.ExcelFile()
                    color_book = excel_file_color.workbook

                    if len(TSD_ids) != 0:
                        units = self.Unit.objects.filter(self.Q(procedureresult__work_order=work_order)&self.Q(procedureresult__test_sequence_definition__id__in=TSD_ids)).distinct()
                    elif len(unit_ids) != 0:
                        units = self.Unit.objects.filter(self.Q(procedureresult__work_order=work_order)&self.Q(id__in=unit_ids)).distinct()
                    else:
                        units = self.Unit.objects.filter(procedureresult__work_order=work_order).distinct()

                    for unit in units:
                        unit_sheet.append([
                            unit.unit_type.manufacturer.name,
                            #unit.unit_type.module_property.module_technology,
                            unit.unit_type.model,
                            unit.serial_number,
                            None,
                            unit.unit_type.module_property.system_voltage,
                            unit.unit_type.module_property.module_width,
                            unit.unit_type.module_property.module_height,
                            unit.unit_type.module_property.nameplate_pmax,
                            unit.unit_type.module_property.voc,
                            unit.unit_type.module_property.vmp,
                            unit.unit_type.module_property.isc,
                            unit.unit_type.module_property.imp
                        ])

                        if len(procedure_ids) != 0:
                            procedure_query = self.ProcedureResult.objects.filter(self.Q(unit=unit)&self.Q(procedure_definition__id__in=procedure_ids)).order_by('test_sequence_definition','linear_execution_group').select_related('unit').prefetch_related('stepresult_set__measurementresult_set')
                        else:
                            procedure_query = self.ProcedureResult.objects.filter(unit=unit).order_by('test_sequence_definition','linear_execution_group').select_related('unit').prefetch_related('stepresult_set__measurementresult_set')

                        for test in procedure_query:
                            if "I-V" in test.procedure_definition.name:
                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [unit.serial_number, test.test_sequence_definition.name, test.name]
                                    for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                        if measurement.measurement_result_type.name == 'result_files':
                                            if measurement.result_files.all().count():
                                                has_result = step_result.measurementresult_set.filter(result_files__name__icontains="Results")
                                                for azurefile in measurement.result_files.all():
                                                    if has_result and "Results" in azurefile.file.name:
                                                        path = "Flash Data/"
                                                        if "200" in step_result.name:
                                                            path += "200W/"
                                                        elif "Rear" in test.procedure_definition.name:
                                                            path += "Rear/"
                                                        path += "FLASH/{}/{}".format(
                                                            test.test_sequence_definition.name,
                                                            test.name
                                                            )
                                                        bytes = azurefile.file.file.read()
                                                        zf.writestr('{}/{}/{}/{}/{}'.format(
                                                            work_order.project.number,
                                                            work_order.name,
                                                            "Flash Data",
                                                            path,
                                                            azurefile.file.name,
                                                            ),
                                                            bytes)
                                                    elif not has_result:
                                                            path = "Flash Data/"
                                                            if "200" in step_result.name:
                                                                path += "200W/"
                                                            elif "Rear" in test.procedure_definition.name:
                                                                path += "Rear/"
                                                            path += "FLASH/{}/{}".format(
                                                                test.test_sequence_definition.name,
                                                                test.name
                                                                )
                                                            bytes = azurefile.file.file.read()
                                                            zf.writestr('{}/{}/{}/{}/{}'.format(
                                                                work_order.project.number,
                                                                work_order.name,
                                                                "Flash Data",
                                                                path,
                                                                azurefile.file.name,
                                                                ),
                                                                bytes)
                                                    else:
                                                        continue
                                        else:
                                            data.append(getattr(measurement, measurement.measurement_result_type.name))
                                    flash_sheet.append(data)
                            elif "Wet Leakage" in test.procedure_definition.name:
                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [unit.serial_number, test.test_sequence_definition.name, test.name]
                                    for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                        data.append(getattr(measurement, measurement.measurement_result_type.name))
                                    wl_sheet.append(data)
                            elif "Visual Inspection" in test.procedure_definition.name:
                                skip = False
                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [unit.serial_number, test.test_sequence_definition.name, test.name]
                                    for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                        if step_result.name == "Inspect Module":
                                            if getattr(measurement, measurement.measurement_result_type.name):
                                                skip = True
                                                data.append("No Defects Observed")
                                                vi_sheet.append(data)
                                        elif skip:
                                            break
                                        else:
                                            if measurement.measurement_result_type.name == 'result_files':
                                                if measurement.result_files.all().count():
                                                    for azurefile in measurement.result_files.all():
                                                        bytes = azurefile.file.file.read()
                                                        zf.writestr('{}/{}/{}/{}'.format(
                                                            work_order.project.number,
                                                            work_order.name,
                                                            "VI Images",
                                                            azurefile.file.name,
                                                            ),
                                                            bytes)
                                                        data.append(azurefile.file.name)
                                            else:
                                                if measurement.result_defect != None and measurement.measurement_result_type.name == "result_defect":
                                                    data.append(measurement.result_defect.short_name)
                                                    data.append(measurement.result_defect.category)
                                                else:
                                                    data.append(getattr(measurement, measurement.measurement_result_type.name))
                                    if skip:
                                        break
                                    vi_sheet.append(data)
                            elif "EL Image" in test.procedure_definition.name:
                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [unit.serial_number, test.test_sequence_definition.name, test.name]
                                    for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                        if measurement.measurement_result_type.name == 'result_files':
                                            if measurement.result_files.all().count():
                                                for azurefile in measurement.result_files.all():
                                                    filetype=''
                                                    if "RAW" in azurefile.file.name:
                                                        continue

                                                    if azurefile.file.name.lower().endswith(('xls','xlsx','txt','csv')):
                                                        filetype = 'DataFiles'
                                                    else:
                                                        filetype = 'ImageFiles'

                                                    if filetype == 'ImageFiles':
                                                        temp = azurefile.file.name.split(".")
                                                        name = temp[0]
                                                        for sbstr in range(1, len(temp) - 1):
                                                            name = name + "." + temp[sbstr]

                                                        name = "{}-{}-{}.{}".format(name, test.test_sequence_definition.name, test.name, temp[len(temp)-1])
                                                    bytes = self.BytesIO(azurefile.file.file.read())
                                                    bytes.seek(0)
                                                    tempImage = self.Image.open(bytes)
                                                    width, height = tempImage.size
                                                    tempImage = tempImage.rotate(-90, expand=True)
                                                    tempImage = tempImage.resize((height, width))
                                                    tempImage = self.ImageOps.grayscale(tempImage)
                                                    tempImage = self.ImageOps.autocontrast(tempImage)

                                                    image_bytes = self.BytesIO()
                                                    tempImage.save(image_bytes, format='jpeg', quality=75)
                                                    tempImage.close()
                                                    image_bytes.seek(0)

                                                    zf.writestr('{}/{}/{}/{}'.format(
                                                        work_order.project.number,
                                                        work_order.name,
                                                        "EL Images",
                                                        name,
                                                        ),
                                                        image_bytes.getvalue())
                                                    image_bytes.close()
                                                    bytes.close()
                                                # continue
                                        else:
                                            data.append(getattr(measurement, measurement.measurement_result_type.name))
                                el_sheet.append(data)
                            elif "Colorimeter" in test.procedure_definition.name:
                                if "{}".format(unit.serial_number) not in color_book.sheetnames:
                                    color_sheet = color_book.create_sheet(title = "{}".format(unit.serial_number))
                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [unit.serial_number, test.test_sequence_definition.name, test.name]
                                    for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                        if step_result.name == "Measure Color" and measurement.measurement_result_type.name == "result_string":
                                            if measurement.result_string == None:
                                                pass
                                                #csv[result.name][result.procedure_definition.name]['body'].append('N/A')
                                            else:
                                                #Convert meassurement string to data values
                                                data = self.json.loads(measurement.result_string)
                                                data = data["values"]

                                                #Header Rows
                                                color_sheet.append([
                                                    "Position",
                                                    "L*",
                                                    "A*",
                                                    "B*"
                                                ])

                                                #Append values to sheet, create columns as we go
                                                l_val, a_val, b_val = [], [], []

                                                color_sheet.append(["Serial Number:", unit.serial_number, "TSD:", test.test_sequence_definition.name, "LEG:", test.name])
                                                for value in data:
                                                    row = [value["position"], value["l_value"], value["a_value"], value["b_value"]]
                                                    l_val.append(value["l_value"])
                                                    a_val.append(value["a_value"])
                                                    b_val.append(value["b_value"])
                                                    color_sheet.append(row)

                                                color_sheet.append(["Average", self.np.average(l_val), self.np.average(a_val), self.np.average(b_val)])

                                                color_sheet.append(["STD", self.np.std(l_val), self.np.std(a_val), self.np.std(b_val)])
                                                color_sheet.append(["", "", "", ""])
                                    #color_sheet.append(data)
                            elif "Diode Test" in test.procedure_definition.name:
                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [unit.serial_number, test.test_sequence_definition.name, test.name]
                                    for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                        data.append(getattr(measurement, measurement.measurement_result_type.name))
                                    dt_sheet.append(data)

                    for sheet in sheets:
                        dims = {}
                        for row in sheet.rows:
                            for cell in row:
                                if cell.value:
                                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                        for col, value in dims.items():
                            sheet.column_dimensions[col].width = value

                    zf.writestr('{}/{}/{}'.format( work_order.project.number, work_order.name, "VI.xlsx"),
                        excel_file_vi.get_memory_file().read())

                    zf.writestr('{}/{}/{}'.format( work_order.project.number, work_order.name, "EL.xlsx"),
                        excel_file_el.get_memory_file().read())

                    zf.writestr('{}/{}/{}'.format( work_order.project.number, work_order.name, "WL.xlsx"),
                        excel_file_wl.get_memory_file().read())

                    zf.writestr('{}/{}/{}'.format( work_order.project.number, work_order.name, "DT.xlsx"),
                        excel_file_dt.get_memory_file().read())

                    zf.writestr('{}/{}/{}'.format( work_order.project.number, work_order.name, "Flash.xlsx"),
                        excel_file_flash.get_memory_file().read())

                    zf.writestr('{}/{}/{}'.format( work_order.project.number, work_order.name, "Colorimeter.xlsx"),
                        excel_file_color.get_memory_file().read())

                    zf.writestr('{}/{}/{}'.format( work_order.project.number, work_order.name, "Units.xlsx"),
                        excel_file_units.get_memory_file().read())


            filename=self.timezone.now().strftime('%b-%d-%Y-%H%M%S')
            response = self.HttpResponse(mem_zip.getvalue(), content_type='application/x-zip-compressed')
            response['Content-Disposition'] = 'attachment; filename={}.zip'.format(filename)
            return response
        return "IDs not found."
