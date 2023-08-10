from lsdb.models import TestSequenceDefinition
from lsdb.models import ProcedureDefinition
from lsdb.models import Group
from lsdb.models import Disposition
from lsdb.models import UnitTypeFamily

from openpyxl import load_workbook

sequencefile = "lsdb/fixtures/loader/LSDB Sequence -Procedure Breakdown.xlsx"
wb = load_workbook(sequencefile)
ws = wb["Test Sequence Detail"]

tc600 = ws['C']


{
    "name": "UV 6.5",
    "description": "",
    "work_in_progress_must_comply": false,
    "group": 4,
    "supersede": false,
    "disposition": 7,
    "version": "0.1",
    "unit_type_family": 1,
    "asset_types": [1],
    "linear_execution_group": 1
}

4 180 Days FE
5 Colorimeter
6 Connector Test
7 DH1000
1 Diode Test
8 Dynamical Mechancial Load Test
9 HF10
10 IAM Measurement
11 IEC 61853-1 Measurement
12 Isc EL Image
14 LIC Flash Test
13 LeTID 162
15 Lightsoak 10kWh/m2
16 Lightsoak 30kWh/m2
17 Lightsoak 40kWh/m2
18 Low Current EL
19 PID 96
21 STC Flash Test
20 Static Mechanical Load
22 TC200
23 TC50
25 UV 6.5
24 UV65
2 Visual Inspection
3 Wet Leakage
{
    "name": "",
    "description": "",
    "notes": "",
    "disposition": 7,
    "version": "0.1",
    "group": 1,
    "unit_type_family": 1
}
Thermal Cycling 600
Damp Heat 2000
Backsheet Durability Sequence
Mechanical Stress Sequence
PotentialInducedDegradation 192
LeTID 486
PAN File
IAM
FE
Light Induced Degradation

Thermal Cycling 600:
[
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 40kWh/m2","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"LIC Flash Test","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 6,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 7,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 8,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 9,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6},
{"execution_group_number": 10,"execution_group_name":"TC200","allow_skip": true,"procedure_definition": 22},
{"execution_group_number": 11,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 12,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 13,"execution_group_name":"Visual ","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 14,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 15,"execution_group_name":"TC200","allow_skip": true,"procedure_definition": 22},
{"execution_group_number": 16,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 17,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 18,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 19,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 20,"execution_group_name":"TC200","allow_skip": true,"procedure_definition": 22},
{"execution_group_number": 21,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 22,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 23,"execution_group_name":"Visual ","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 24,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 25,"execution_group_name":"Diode Test","allow_skip": true,"procedure_definition": 1},
{"execution_group_number": 26,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6}
]
Damp Heat 2000
[
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 40kWh/m2","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"LIC Flash Test","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 6,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 7,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 8,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 9,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6},
{"execution_group_number": 10,"execution_group_name":"Low Current EL","allow_skip": true,"procedure_definition": 18},
{"execution_group_number": 11,"execution_group_name":"DH1000","allow_skip": true,"procedure_definition": 7},
{"execution_group_number": 12,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 13,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 14,"execution_group_name":"Visual ","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 15,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 16,"execution_group_name":"DH1000","allow_skip": true,"procedure_definition": 7},
{"execution_group_number": 17,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 18,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 19,"execution_group_name":"Visual ","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 20,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 21,"execution_group_name":"Diode Test","allow_skip": true,"procedure_definition": 1},
{"execution_group_number": 22,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6}
]
Backsheet Durability Sequence
[
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 4,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 5,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5},
{"execution_group_number": 6,"execution_group_name":"DH1000","allow_skip": true,"procedure_definition": 7},
{"execution_group_number": 7,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 8,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 9,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 10,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5},
{"execution_group_number": 11,"execution_group_name":"UV65","allow_skip": true,"procedure_definition": 24},
{"execution_group_number": 12,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 13,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5},
{"execution_group_number": 14,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 15,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 16,"execution_group_name":"TC50","allow_skip": true,"procedure_definition": 23},
{"execution_group_number": 17,"execution_group_name":"HF10","allow_skip": true,"procedure_definition": 9},
{"execution_group_number": 18,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 19,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 20,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 21,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5},
{"execution_group_number": 22,"execution_group_name":"UV65","allow_skip": true,"procedure_definition": 24},
{"execution_group_number": 23,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 24,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 25,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 26,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5},
{"execution_group_number": 27,"execution_group_name":"TC50","allow_skip": true,"procedure_definition": 23},
{"execution_group_number": 28,"execution_group_name":"HF10","allow_skip": true,"procedure_definition": 9},
{"execution_group_number": 29,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 30,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 31,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 32,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5},
{"execution_group_number": 33,"execution_group_name":"UV65","allow_skip": true,"procedure_definition": 24},
{"execution_group_number": 34,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 35,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 36,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 37,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5},
{"execution_group_number": 38,"execution_group_name":"TC50","allow_skip": true,"procedure_definition": 23},
{"execution_group_number": 39,"execution_group_name":"HF10","allow_skip": true,"procedure_definition": 9},
{"execution_group_number": 40,"execution_group_name":"UV 6.5","allow_skip": true,"procedure_definition": 25},
{"execution_group_number": 41,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 42,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 43,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 44,"execution_group_name":"Colorimeter","allow_skip": true,"procedure_definition": 5}
]
Mechanical Stress Sequence
[
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 30kWh/m2","allow_skip": true,"procedure_definition": 16},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 6,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 7,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 8,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 9,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 10,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 11,"execution_group_name":"LIC Flash Test","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 12,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 13,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 14,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 15,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6},
{"execution_group_number": 16,"execution_group_name":"Static Mechanical Load","allow_skip": true,"procedure_definition": 20},
{"execution_group_number": 17,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 18,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 19,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 20,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 21,"execution_group_name":"Dynamical Mechancial Load Test","allow_skip": true,"procedure_definition":  8},
{"execution_group_number": 22,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 23,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 24,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 25,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 26,"execution_group_name":"TC50","allow_skip": true,"procedure_definition": 23},
{"execution_group_number": 27,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 28,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 29,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 30,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 31,"execution_group_name":"HF10","allow_skip": true,"procedure_definition":  9},
{"execution_group_number": 32,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 33,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 34,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 35,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 36,"execution_group_name":"Diode Test","allow_skip": true,"procedure_definition": 1},
{"execution_group_number": 37,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6}
]
PID 192 [
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 30kWh/m2","allow_skip": true,"procedure_definition": 16},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 6,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 7,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 8,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 9,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 10,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 11,"execution_group_name":"LIC Flash Test","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 12,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 13,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 14,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 15,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6},
{"execution_group_number": 16,"execution_group_name":"Low Current EL","allow_skip": true,"procedure_definition": 18},
{"execution_group_number": 17,"execution_group_name":"PID 96","allow_skip": true,"procedure_definition":19},
{"execution_group_number": 18,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 19,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 20,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 21,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 22,"execution_group_name":"PID 96","allow_skip": true,"procedure_definition":19},
{"execution_group_number": 23,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 24,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 25,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 26,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 27,"execution_group_name":"Diode Test","allow_skip": true,"procedure_definition": 1},
{"execution_group_number": 28,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6},
{"execution_group_number": 29,"execution_group_name":"Low Current EL","allow_skip": true,"procedure_definition": 18}]

LeTID 486[
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 30kWh/m2","allow_skip": true,"procedure_definition": 16},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 6,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 7,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 8,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 9,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 10,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 11,"execution_group_name":"LIC Flash Test","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 12,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 13,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 14,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 15,"execution_group_name":"Connector Test","allow_skip": true,"procedure_definition": 6},
{"execution_group_number": 16,"execution_group_name":"Low Current EL","allow_skip": true,"procedure_definition": 18},
{"execution_group_number": 17,"execution_group_name":"LeTID 162","allow_skip": true,"procedure_definition":  13},
{"execution_group_number": 18,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 19,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 20,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 21,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 22,"execution_group_name":"LeTID 162","allow_skip": true,"procedure_definition":  13},
{"execution_group_number": 23,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 24,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 25,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 26,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 27,"execution_group_name":"LeTID 162","allow_skip": true,"procedure_definition":  13},
{"execution_group_number": 28,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 29,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 30,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 31,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 32,"execution_group_name":"Low Current EL","allow_skip": true,"procedure_definition": 18}]

]
PAN File
[
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 30kWh/m2","allow_skip": true,"procedure_definition": 16},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 6,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 7,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 8,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 9,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 10,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 11,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 12,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 13,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 14,"execution_group_name":"IEC 61853-1 Measurement","allow_skip": true,"procedure_definition": 11}
]
IAM
[
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 30kWh/m2","allow_skip": true,"procedure_definition": 16},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 6,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 7,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 8,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 9,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 10,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 11,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 12,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 13,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 14,"execution_group_name":"IAM Measurement","allow_skip": true,"procedure_definition": 10}
]
FE [
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 4,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 5,"execution_group_name":"180 Days FE","allow_skip": true,"procedure_definition": 4},
{"execution_group_number": 6,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 7,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 8,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 9,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3},
{"execution_group_number": 10,"execution_group_name":"180 Days FE","allow_skip": true,"procedure_definition": 4},
{"execution_group_number": 11,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 12,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 13,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 14,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3}]
LID [
{"execution_group_number": 1,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 2,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 3,"execution_group_name":"Lightsoak 30kWh/m2","allow_skip": true,"procedure_definition": 16},
{"execution_group_number": 4,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 5,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 6,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 7,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 8,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 9,"execution_group_name":"Lightsoak 10kWh/m2","allow_skip": true,"procedure_definition": 15},
{"execution_group_number": 10,"execution_group_name":"STC Flash Test","allow_skip": true,"procedure_definition": 21},
{"execution_group_number": 11,"execution_group_name":"LIC Flash Test","allow_skip": true,"procedure_definition": 14},
{"execution_group_number": 12,"execution_group_name":"Isc EL Image","allow_skip": true,"procedure_definition": 12},
{"execution_group_number": 13,"execution_group_name":"Visual","allow_skip": true,"procedure_definition": 2},
{"execution_group_number": 14,"execution_group_name":"Wet Leakage","allow_skip": true,"procedure_definition": 3}
]
