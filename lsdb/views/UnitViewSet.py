import json
import pandas as pd
from io import BytesIO
from django.db import IntegrityError, transaction
from django_filters import rest_framework as filters
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import F, Q, Max, Min
from django.db.models.functions import Coalesce
# from openpyxl.writer.excel import save_virtual_workbook
# from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, NamedStyle, Side, PatternFill, Font, GradientFill, Alignment

from rest_framework import viewsets
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework_tracking.mixins import LoggingMixin
from rest_framework.status import (HTTP_400_BAD_REQUEST)
from rest_framework.permissions import AllowAny

from lsdb.serializers import AzureFileSerializer
from lsdb.serializers import AssetSerializer
from lsdb.serializers import DispositionCodeListSerializer
from lsdb.serializers import NoteSerializer
from lsdb.serializers import ProcedureResultSerializer
from lsdb.serializers import UnitSerializer
from lsdb.serializers import UnitGroupedTravelerSerializer
from lsdb.serializers import UnitGroupedAssetTypeSerializer
from lsdb.serializers import UnitTravelerSerializer
from lsdb.serializers import UnitDumpSerializer

from lsdb.models import Asset
from lsdb.models import AssetType
from lsdb.models import AzureFile
from lsdb.models import Customer
from lsdb.models import Disposition
from lsdb.models import DispositionCode
from lsdb.models import MeasurementResult
from lsdb.models import Note
from lsdb.models import ProcedureResult
from lsdb.models import StepResult
from lsdb.models import TestSequenceDefinition
from lsdb.models import Unit
from lsdb.models import UnitType
from django.contrib.auth.models import User

from lsdb.permissions import ConfiguredPermission
# from lsdb.utils.NoteUtils import create_note
from lsdb.utils.Crypto import encrypt, decrypt
from lsdb.utils import ExcelFile


class UnitFilter(filters.FilterSet):
    # projects =

    class Meta:
        model = Unit
        fields = {
            'serial_number': ['exact', 'icontains'],
            # 'project_set',
        }


class UnitViewSet(LoggingMixin, viewsets.ModelViewSet):
    """
    API endpoint that allows Unit to be viewed or edited.
    """
    logging_methods = ['POST', 'PUT', 'PATCH', 'DELETE']
    queryset = Unit.objects.all()
    serializer_class = UnitSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = UnitFilter
    permission_classes = [ConfiguredPermission]

    # Override _clean_data to decode data with ignore instead of replace to ignore
    # errors in decode so when the logger inserts the data to db, it will not hit
    # any decoding/encoding issues
    # def _clean_data(self, data):
    #     if isinstance(data, bytes):
    #         # data = data.decode(errors='ignore')
    #         data = 'CLEANED FILE DATA'
    #     return super(NoteViewSet, self)._clean_data(data)

    @action(detail=False, methods=['get', ],
            serializer_class=DispositionCodeListSerializer,
            )
    def dispositions(self, request, pk=None):
        self.context = {'request': request}
        serializer = DispositionCodeListSerializer(DispositionCode.objects.get(
            name='units'),
            many=False,
            context={'request': request})
        return Response(serializer.data)

    # @transaction.atomic
    # @action(detail=True, methods=['get','post'],
    #     permission_classes=(ConfiguredPermission,),
    #     serializer_class = UnitSerializer,
    # )
    # def add_note(self, request, pk=None):
    #     """
    #     This action is used to add notes to a unit. This also adds the note to the Unit's Project
    #     POST:
    #     {
    #         "subject": "Unit note subject",
    #         "text": "Unit note text body",
    #         "type":$ID,
    #         "owner": $ID,
    #         "disposition": $ID,
    #         "labels": [$ID1, $ID2...],
    #         "groups": [$ID1, $ID2...],
    #         "tagged_users": [$ID1, $ID2...]
    #     }
    #     """
    #     self.context = {'request':request}
    #     unit = Unit.objects.get(id=pk)
    #     if request.method == "POST":
    #         params = json.loads(request.body)
    #         noteParams = {
    #             "model" : "unit",
    #             "pk" : pk,
    #             "user" : request.user,
    #             "subject" : params.get('subject'),
    #             "text" : params.get('text'),
    #             "note_type" : params.get('type'),
    #             }
    #         if params.get('owner') and params.get('owner') != -1:
    #             noteParams["owner"] = User.objects.get(id=params.get('owner'))
    #         if params.get('disposition') and params.get('disposition')!= -1 :
    #             noteParams["disposition"] =Disposition.objects.get(id=params.get('disposition'))
    #         newNote = create_note(
    #             **noteParams
    #         )
    #         # Unit Notes need to trickle notes up to Project:
    #         for project in unit.project_set.all():
    #             project.notes.add(newNote)
    #         newNote.labels.set(params.get('labels'))
    #         newNote.groups.set(params.get('groups'))
    #         newNote.tagged_users.set(params.get('tagged_users'))
    #         newNote.save()
    #         serializer = NoteSerializer(newNote, many=False, context=self.context)
    #     else:
    #         serializer = NoteSerializer([], many=True, context=self.context)
    #     return Response(serializer.data)

    @transaction.atomic
    @action(detail=True, methods=['get', 'post'],
            permission_classes=(ConfiguredPermission,),
            serializer_class=UnitSerializer,
            )
    def append_test_sequence(self, request, pk=None):
        """
        This action is used to add a test sequence to a unit's procedure result list.
        POST:
        {
        "test_sequence_definition":$ID
        }
        """
        self.context = {'request': request}
        unit = Unit.objects.get(id=pk)

        if request.method == "POST":
            params = json.loads(request.body)
            try:
                new_tsd = TestSequenceDefinition.objects.get(id=params.get('test_sequence_definition'))
            except:
                return Response({
                    'message': 'test_sequence with ID:{} not found. Contact engineering.'.format(
                        params.get('test_sequence_definition'))
                }, status=HTTP_400_BAD_REQUEST)
            # KLUGE this process of picking the work_order from the first procedure is unsustainable
            work_order = unit.procedureresult_set.first().work_order
            last_leg = unit.procedureresult_set.last().linear_execution_group
            # start adding new procedures with a new LEG starting point
            procedures = []
            # KLUGE this code is copied all over, need to refactor and modularize it for general use and optimization
            for execution in new_tsd.procedureexecutionorder_set.all():
                procedure_result = ProcedureResult.objects.create(
                    unit=unit,
                    name=execution.execution_group_name + ' (APPENDED)',
                    disposition=None,
                    group=execution.procedure_definition.group,
                    work_order=work_order,
                    procedure_definition=execution.procedure_definition,
                    version=execution.procedure_definition.version,
                    linear_execution_group=execution.execution_group_number + last_leg,
                    test_sequence_definition=new_tsd,
                    allow_skip=execution.allow_skip,
                )
                for step_execution in execution.procedure_definition.stepexecutionorder_set.all():
                    step_result = StepResult.objects.create(
                        name=step_execution.execution_group_name,
                        procedure_result=procedure_result,
                        step_definition=step_execution.step_definition,
                        execution_number=0,
                        disposition=None,
                        start_datetime=None,
                        duration=0,
                        test_step_result=None,
                        archived=False,
                        description=None,
                        step_number=0,
                        step_type=step_execution.step_definition.step_type,
                        linear_execution_group=step_execution.execution_group_number,
                        allow_skip=step_execution.allow_skip,
                    )
                    for measurement_definition in step_execution.step_definition.measurementdefinition_set.all():
                        measurement_result = MeasurementResult.objects.create(
                            step_result=step_result,
                            measurement_definition=measurement_definition,
                            software_revision=0.0,
                            disposition=None,
                            limit=measurement_definition.limit,
                            station=0,
                            name=measurement_definition.name,
                            record_only=measurement_definition.record_only,
                            allow_skip=measurement_definition.allow_skip,
                            requires_review=measurement_definition.requires_review,
                            measurement_type=measurement_definition.measurement_type,
                            order=measurement_definition.order,
                            report_order=measurement_definition.report_order,
                            measurement_result_type=measurement_definition.measurement_result_type,
                        )
                procedures.append(procedure_result)
            serializer = ProcedureResultSerializer(procedures, many=True, context=self.context)
        else:
            # GET
            serializer = UnitSerializer(unit, many=False, context=self.context)
        return Response(serializer.data)

    @action(detail=False, methods=['get', 'post'],
            serializer_class=UnitDumpSerializer,
            )
    def dump(self, request, pk=None):
        # This will spit out a csv of the whole module serial number list
        self.context = {'request': request}
        units = Unit.objects.all()
        serializer = UnitDumpSerializer(units, many=True, context=self.context)
        return Response(serializer.data)
        # report=[]
        # for unit in units:
        #     report.append([unit.id, unit.serial_number, unit.unit_type.manufacturer.name, unit.unit_type.model])
        # return Response(report)

    @transaction.atomic
    @action(detail=True, methods=['get'],
            serializer_class=UnitTravelerSerializer,
            )
    def measurement_history(self, request, pk=None):
        self.context = {'request': request}
        # queryset = Unit.objects.get(id=pk)#.prefetch_related('step_results__measurement_results')
        queryset = Unit.objects.prefetch_related(
            'procedureresult_set__stepresult_set__measurementresult_set').get(id=pk)
        serializer = self.serializer_class(queryset, many=False, context=self.context)
        # print(serializer.data)
        return Response(serializer.data)

    @transaction.atomic
    @action(detail=True, methods=['get'],
            serializer_class=UnitGroupedTravelerSerializer,
            )
    def grouped_history(self, request, pk=None):
        self.context = {'request': request}
        file = request.query_params.get('file', 'dummy')
        # queryset = Unit.objects.get(id=pk)#.prefetch_related('step_results__measurement_results')
        queryset = Unit.objects.get(id=pk)
        serializer = self.serializer_class(queryset, many=False, context=self.context)
        if file.upper() == 'EXCEL':
            # wb = Workbook()
            myFile = ExcelFile()
            wb = myFile.workbook

            # Styles:
            hairline = Side(border_style="hair", color="000000")
            thin = Side(border_style="thin", color="000000")
            medium = Side(border_style="medium", color="000000")
            thick = Side(border_style='thick', color="000000")

            center_b = NamedStyle(name='center_b')
            center_b.border = Border(start=medium, end=medium, outline=True)
            center_b.fill = PatternFill("solid", fgColor="00FFFFFF")
            center_b.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=False, shrink_to_fit=True)
            center_b.font = Font(name='Arial', b=True, size=10)

            center_bl = NamedStyle(name='center_bl')
            center_bl.border = Border(start=medium, end=medium, outline=True, top=medium, bottom=medium)
            center_bl.alignment = Alignment(horizontal="center", vertical="center")
            center_bl.font = Font(name='Arial', b=True, size=22)

            center_bxl = NamedStyle(name='center_bxl')
            center_bxl.border = Border(start=medium, end=medium, outline=True, top=medium, bottom=medium)
            center_bxl.alignment = Alignment(horizontal="center", vertical="center",
                                             wrap_text=False,
                                             shrink_to_fit=True)
            center_bxl.font = Font(name='Arial', b=True, size=26)

            center = NamedStyle(name='center')
            center.border = Border(start=medium, end=medium, top=hairline, bottom=hairline, outline=True)
            center.alignment = Alignment(horizontal="center", vertical="center",
                                         wrap_text=False, shrink_to_fit=True)
            center.font = Font(name='Arial', b=False, size=10)

            leg_num = NamedStyle(name='leg_num')
            leg_num.alignment = Alignment(horizontal="center", vertical="center")
            # leg_num.border = Border(top=medium, left=medium, right=medium, bottom=medium)
            leg_num.border = Border(top=medium, left=medium, right=medium, bottom=medium, outline=True)
            leg_num.fill = PatternFill("solid", fgColor="00FFFF00")
            leg_num.font = Font(name='Arial', b=True, size=24, color="00000000")

            leg_head = NamedStyle(name='leg_head')
            leg_head.alignment = Alignment(horizontal="center", vertical="center")
            leg_head.border = Border(top=medium, left=medium, right=medium, bottom=medium, outline=True)
            leg_head.fill = PatternFill("solid", fgColor="00FFFF00")
            leg_head.font = Font(name='Arial', b=True, size=12, color="00000000")

            test_title = NamedStyle(name='test_title')
            test_title.border = Border(left=medium, right=medium, top=hairline, bottom=hairline, outline=True)
            test_title.alignment = Alignment(horizontal="left", vertical="center",
                                             wrap_text=False,
                                             shrink_to_fit=True)
            test_title.font = Font(name='Arial', b=False, size=11)

            tiny_grey = NamedStyle(name='tiny_grey')
            tiny_grey.border = Border(left=hairline, right=hairline, outline=True)
            tiny_grey.fill = PatternFill("solid", fgColor="00DCE6F2")
            tiny_grey.alignment = Alignment(horizontal="center", vertical="center",
                                            wrap_text=False,
                                            shrink_to_fit=True)
            tiny_grey.font = Font(name='Arial', b=False, size=8)

            test_data = NamedStyle(name='test_data')
            test_data.border = Border(left=hairline, right=hairline, top=hairline, bottom=hairline, outline=True)
            test_data.fill = PatternFill("solid", fgColor="00FFFFFF")
            test_data.alignment = Alignment(horizontal="center", vertical="center",
                                            wrap_text=False,
                                            shrink_to_fit=True)
            test_data.font = Font(name='Arial', b=False, size=8)

            wb.add_named_style(center_b)
            wb.add_named_style(center_bl)
            wb.add_named_style(center_bxl)
            wb.add_named_style(center)
            wb.add_named_style(leg_num)
            wb.add_named_style(leg_head)
            wb.add_named_style(test_title)
            wb.add_named_style(tiny_grey)
            wb.add_named_style(test_data)

            sheet = wb.active
            sheet.page_setup.fitToWidth = 1

            unit_type = serializer.data.pop('unit_type')
            calibration_results = serializer.data.pop('calibration_results')
            sequences_results = serializer.data.pop('sequences_results')
            module_property = unit_type.pop('module_property')
            sheet.merge_cells('A1:D2')
            sheet.merge_cells('E1:Q2')
            for i in range(1, 28):
                sheet.column_dimensions[get_column_letter(i)].width = 3.5
            if module_property.get('bifacial'):
                sheet['A1'] = 'Bifacial'
                sheet['A1'].style = 'center_bl'
            sheet['E1'] = serializer.data.pop('project_number')
            sheet['E1'].style = 'center_bl'

            for row in range(3, 9):
                for col in [[1, 4], [5, 8], [9, 12], [13, 17]]:
                    sheet.merge_cells(start_column=col[0], end_column=col[1], start_row=row, end_row=row)

            sheet.cell(row=3, column=1, value='Start Date').style = 'center_b'
            sheet.cell(row=3, column=5, value=serializer.data.pop('start_datetime')).style = 'center'
            sheet.cell(row=3, column=9, value='Mpp').style = 'center_b'
            sheet.cell(row=3, column=13, value=module_property.get('nameplate_pmax')).style = 'center'

            sheet.cell(row=4, column=1, value='Manufacturer').style = 'center_b'
            sheet.cell(row=4, column=5, value=unit_type.get('manufacturer_name')).style = 'center'
            sheet.cell(row=4, column=9, value='Voc').style = 'center_b'
            sheet.cell(row=4, column=13, value=module_property.get('voc')).style = 'center'

            sheet.cell(row=5, column=1, value='Module Type').style = 'center_b'
            sheet.cell(row=5, column=5, value=module_property.get('module_technology_name')).style = 'center'
            sheet.cell(row=5, column=9, value='Isc').style = 'center_b'
            sheet.cell(row=5, column=13, value=module_property.get('isc')).style = 'center'

            sheet.cell(row=6, column=1, value='Project Manager').style = 'center_b'
            sheet.cell(row=6, column=5, value=serializer.data.pop('project_manager')).style = 'center'
            sheet.cell(row=6, column=9, value='Vmp').style = 'center_b'
            sheet.cell(row=6, column=13, value=module_property.get('vmp')).style = 'center'

            sheet.cell(row=7, column=1, value='Calibration').style = 'center_b'
            sheet.cell(row=7, column=5, value='?????').style = 'center'
            sheet.cell(row=7, column=9, value='Imp').style = 'center_b'
            sheet.cell(row=7, column=13, value=module_property.get('imp')).style = 'center'

            sheet.cell(row=8, column=1, value='Dimensions').style = 'center_b'
            sheet.cell(row=8, column=5, value='{:n}x{:n}mm'.format(module_property.get('module_height'),
                                                                   module_property.get(
                                                                       'module_width'))).style = 'center'
            sheet.cell(row=8, column=9, value='Voltage').style = 'center_b'
            sheet.cell(row=8, column=13, value='{}V'.format(module_property.get('system_voltage'))).style = 'center'

            sheet.merge_cells(start_column=1, end_column=12, start_row=9, end_row=10)
            sheet.merge_cells(start_column=13, end_column=24, start_row=9, end_row=10)
            sheet.cell(row=9, column=1, value=serializer.data.pop('customer_name')).style = 'center_bxl'
            sheet.cell(row=9, column=13, value=serializer.data.pop('work_order_name')).style = 'center_bxl'

            sheet.merge_cells('R1:X2')
            sheet['R1'] = serializer.data.pop('serial_number')
            sheet['R1'].style = 'center_bxl'

            sheet.merge_cells('R3:X8')
            sheet['R3'] = serializer.data.pop('test_sequence_definition_name')
            sheet['R3'].style = 'center_bxl'
            sheet['R3'].alignment = Alignment(horizontal="center", vertical="center",
                                              wrap_text=True)
            sheet['R3'].fill = PatternFill("solid", fgColor="00003300")
            sheet['R3'].font = Font(name='Arial', b=True, size=26, color="00FFFFFF")

            # Apply heading styles:
            for row in range(3, 9):
                for col in ['A', 'L']:
                    cell = '{}{}'.format(col, row)
                    sheet[cell].style = 'center_b'
                for col in ['F', 'P']:
                    cell = '{}{}'.format(col, row)
                    sheet[cell].style = 'center'

            current_row = 11
            for sequences in sequences_results:
                sheet.merge_cells('A{}:B{}'.format(current_row, current_row + 1))
                sheet.merge_cells('C{}:X{}'.format(current_row, current_row))
                sheet['A{}'.format(current_row)] = sequences.get('linear_execution_group')
                sheet['A{}'.format(current_row)].style = 'leg_num'
                # leg_num = sequences.get('linear_execution_group')
                sheet['C{}'.format(current_row)] = sequences.get('name')
                sheet['C{}'.format(current_row)].style = 'leg_head'
                current_row += 1
                for col in range(0, 5):
                    sheet.merge_cells(start_column=5 + (col * 2), end_column=6 + (col * 2), start_row=current_row,
                                      end_row=current_row)

                sheet.cell(row=current_row, column=5, value='Date').style = 'tiny_grey'
                sheet.cell(row=current_row, column=7, value='Initials').style = 'tiny_grey'
                sheet.cell(row=current_row, column=9, value='Review Date').style = 'tiny_grey'
                sheet.cell(row=current_row, column=11, value='Reviewer').style = 'tiny_grey'
                sheet.cell(row=current_row, column=13, value='Result').style = 'tiny_grey'

                # leg_head.style = 'leg_head'
                # leg_head = sequences.get('name')
                start_row = current_row
                current_row += 1
                for result in sequences.get('procedure_results'):
                    sheet.merge_cells('A{}:D{}'.format(current_row, current_row))
                    sheet['A{}'.format(current_row)].style = 'test_title'
                    sheet['A{}'.format(current_row)] = result.get('procedure_definition_name')
                    for col in range(0, 5):
                        sheet.merge_cells(start_column=5 + (col * 2), end_column=6 + (col * 2), start_row=current_row,
                                          end_row=current_row)

                    sheet.cell(row=current_row, column=5, value=result.get('completion_date')).style = 'test_data'
                    sheet.cell(row=current_row, column=7, value=result.get('username')).style = 'test_data'
                    sheet.cell(row=current_row, column=9, value=result.get('review_datetime')).style = 'test_data'
                    sheet.cell(row=current_row, column=11, value=result.get('reviewed_by_user')).style = 'test_data'
                    sheet.cell(row=current_row, column=13, value=result.get('disposition_name')).style = 'test_data'

                    current_row += 1
                    if result.get('visualizer') == 'stress':
                        # need to add metadata for stressor rows here
                        steps = StepResult.objects.filter(procedure_result__id=result.get('id')).order_by(
                            'linear_execution_group')
                        # steps = steps.exclude(name__iexact='test end')
                        # steps = steps.exclude(name__iexact='test start')
                        for step in steps:
                            sheet.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=4)
                            sheet.cell(row=current_row, column=1, value=step.name).style = 'test_title'
                            for col in range(0, 5):
                                sheet.merge_cells(start_column=5 + (col * 2), end_column=6 + (col * 2),
                                                  start_row=current_row, end_row=current_row)
                            current_row += 1
                sheet.merge_cells(start_column=15, end_column=24, start_row=start_row, end_row=current_row - 1)
                sheet.cell(row=start_row, column=15, value=' ').style = 'center_b'

            # mem_file = BytesIO(save_virtual_workbook(wb))
            filename = timezone.now().strftime('%b-%d-%Y-%H%M%S')
            # response = HttpResponse(mem_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(filename)
            # return response
            return myFile.get_response(filename)
        else:
            return Response(serializer.data)

    @transaction.atomic
    def get_queue(self, group=None, asset=None):
        queryset = ProcedureResult.objects.filter(disposition__isnull=True,
                                                  unit__disposition__complete=False,)\
            .exclude(unit__disposition__name__iexact="in progress") \
            .exclude(work_order__project__disposition__complete=True) \
            .exclude(test_sequence_definition__group__name__iexact="control") \
            .annotate(last_action_date=Coalesce(
                Max('unit__procedureresult__stepresult__measurementresult__date_time'),timezone.now())) \
            .annotate(done_to=Coalesce(Max('unit__procedureresult__linear_execution_group',
                    filter=Q(unit__procedureresult__disposition__isnull=False)),
                    Min('unit__procedureresult__linear_execution_group'))) \
            .distinct()
            # .exclude(unit__procedureresult__work_in_progress_must_comply=True) \
        if asset:
            # if any of asset.asset_types in proceduredefinition.asset_types:
            # print(asset.location.id, queryset)
            queryset = queryset.filter(procedure_definition__asset_types__in=asset.asset_types.all(),
                unit__location__id = asset.location.id,
            )
            # print(queryset)
        # if group:
        #     queryset = queryset.filter(group__name__iexact=group)

        master_data_frame = pd.DataFrame(list(queryset.values(
            'unit__serial_number',
            'test_sequence_definition__name',
            'done_to',
            # 'highest_leg',
            'linear_execution_group',
            'procedure_definition__name',
            'work_order__project__customer__name',
            'work_order__project__number',
            'work_order__name',
            'name',
            'allow_skip',
            'last_action_date',
            'group__name',
            'unit__location__name',
        )))
        master_data_frame['last_action_days'] = (timezone.now() - master_data_frame.last_action_date).astype(
            "timedelta64[D]")
        master_data_frame.dropna(inplace=True)
        # Select everything in the current LEG or above
        filtered = master_data_frame[master_data_frame.linear_execution_group >= master_data_frame.done_to]
        # select only records where LEG <= next highest unskippable
        gframe = filtered.groupby(['unit__serial_number','allow_skip'])
        #filtered = filtered[filtered.allow_skip == False]
        filterframe = gframe[['unit__serial_number','allow_skip','linear_execution_group']].min()
        filterframe = filterframe[filterframe.allow_skip == False]
        filterframe.columns=['f_serial','f_allow_skip','highest_leg']
        filterframe.set_index(['f_serial'], inplace=True)
        filtered = filtered.merge(filterframe, left_on='unit__serial_number', right_on='f_serial', suffixes=(False,False))
        filtered = filtered[filtered.linear_execution_group <= filtered.highest_leg]
        # Now we have the map, fiter down to our char group:
        if group:
            filtered = filtered[filtered.group__name.str.lower() == group.lower()]

        filtered = filtered[[
            'unit__serial_number',
            'test_sequence_definition__name',
            'linear_execution_group',
            'procedure_definition__name',
            'work_order__project__customer__name',
            'work_order__project__number',
            'work_order__name',
            'name',
            'allow_skip',
            'last_action_date',
            'last_action_days',
            'unit__location__name',
        ]]
        filtered.columns = ['serial_number', 'test_sequence', 'linear_execution_group',
                         'procedure_definition', 'customer', 'project_number', 'work_order',
                         'characterization', 'allow_skip',
                         'last_action_date',
                         'last_action_days',
                         'location',
                         ]
        # Excel needs TZ stripped
        filtered.loc[:, ('last_action_date')] = filtered.loc[:, ('last_action_date')].dt.tz_localize(None)
        grouped = filtered.groupby('procedure_definition')
        full = {}
        for name, group in grouped:
            full[name] = group.to_dict(orient='records')
        # serializer = self.serializer_class(queryset, many=False, context=self.context)
        # print(serializer.data)
        return full, filtered

    @transaction.atomic
    @action(detail=False, methods=['get'],
            # serializer_class=UnitGroupedAssetTypeSerializer,
            )
    def characterization_queue(self, request, pk=None):
        # This is the list of "characterization ready" modules grouped by their test type
        self.context = {'request': request}
        # get the asset from the request, pass it to get_queue:
        if request.method=='GET':
            # print('foo',request.GET.get('asset_name',None))
            try:
                asset = Asset.objects.get(name = request.GET.get('asset_name',None))
            except:
                # print('failed')
                asset = None
        full, fcols = self.get_queue('characterizations', asset)
        return Response(full)

    @transaction.atomic
    @action(detail=False, methods=['get'],
            # serializer_class=UnitGroupedAssetTypeSerializer,
            )
    def stress_queue(self, request, pk=None):
        # This is the list of "chamber ready" modules grouped by their stressor
        self.context = {'request': request}
        full, fcols = self.get_queue('stressors')
        return Response(full)

    @transaction.atomic
    @action(detail=False, methods=['get'], serializer_class=UnitSerializer, )
    def end_of_life(self, request):
        self.context = {'request': request}
        units = Unit.objects.annotate(max_leg=Max('procedureresult__linear_execution_group')) \
            .filter(max_leg__isnull=False)

        unit_ids = []

        for unit in units:
            max_leg = unit.max_leg
            if unit.procedureresult_set.filter(linear_execution_group=max_leg, disposition__complete=True):
                unit_ids.append(unit.id)

        final_units = Unit.objects.filter(id__in=unit_ids)

        units_data_frame = pd.DataFrame(list(
            final_units.values('serial_number',
                               'workorder__project__customer__name',
                               'workorder__project__number',
                               'workorder__name',
                               'procedureresult__test_sequence_definition__name',
                               'procedureresult__end_datetime',
                               'procedureresult__linear_execution_group'))
        )

        units_data_frame.columns = [
            'serial_number',
            'customer_name',
            'project_number',
            'work_order_name',
            'test_sequence_definition_name',
            'completion_date',
            'linear_execution_group',
        ]

        units_data_frame.sort_values(by='linear_execution_group', ascending=False, inplace=True)
        units_data_frame.drop_duplicates(subset='serial_number', inplace=True)
        units_data_frame.drop(labels='linear_execution_group', axis=1, inplace=True)

        return Response(units_data_frame.to_dict(orient='records'))

    @transaction.atomic
    @action(detail=False, methods=['get'],
            # serializer_class=UnitGroupedAssetTypeSerializer,
            )
    def in_progress(self, request, pk=None):
        # This is the list of modules currently "in progress" of a stress, grouped by their stressor
        self.context = {'request': request}
        disposition = Disposition.objects.get(name__iexact='in progress')
        queryset = ProcedureResult.objects.filter(
            disposition=disposition,
            stepresult__disposition__isnull=False,
            stepresult__name__iexact="test start",  # KLUGE!
            group__name__iexact='stressors') \
            .exclude(test_sequence_definition__group__name__iexact="control") \
            .distinct()

        master_data_frame = pd.DataFrame(list(queryset.values(
            'unit__serial_number',
            'procedure_definition__asset_types__name',
            'stepresult__measurementresult__asset__name',
            'stepresult__measurementresult__asset__id',
            'test_sequence_definition__name',
            'linear_execution_group',
            'procedure_definition__name',
            'start_datetime',
            'procedure_definition__aggregate_duration'
        )))
        # master_data_frame
        filled = master_data_frame.fillna('foo')  # this needs removal
        filled['procedure_definition__aggregate_duration'] = pd.to_timedelta(
            filled.procedure_definition__aggregate_duration, unit='m')
        filled['eta_datetime'] = filled['start_datetime'] + filled['procedure_definition__aggregate_duration']
        # print(filled)
        filled.columns = [
            'serial_number',
            'asset_type',
            'asset_name',
            'asset_id',
            'test_sequence',
            'linear_execution_group',
            'procedure_definition',
            'start_datetime',
            'procedure_duration',
            'eta_datetime', ]

        filled.sort_values(by=['asset_id', 'start_datetime'], inplace=True)
        filled.drop_duplicates(subset='asset_id', keep='first', inplace=True)
        grouped = filled.groupby('asset_type')

        # filtered = master_data_frame.duplicated(subset='asset_id',keep='first').sum()
        # grouped = filtered.groupby('asset_type')
        full = {}
        for name, group in grouped:
            full[name] = group.to_dict(orient='records')
        # serializer = self.serializer_class(queryset, many=False, context=self.context)
        # print(serializer.data)
        # print(full)
        return Response(full)
        # return Response(serializer.data)

    @transaction.atomic
    @action(detail=False, methods=['get'])
    def todo_queue(self, request, pk=None):
        # This is the list of "chamber ready" modules grouped by their stressor
        full, fcols = self.get_queue()
        mem_file = BytesIO()
        with pd.ExcelWriter(mem_file, engine='openpyxl') as writer:
            fcols.to_excel(writer)
        mem_file.seek(0)
        filename = timezone.now().strftime('%b-%d-%Y-%H%M%S')
        response = HttpResponse(mem_file,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(filename)
        return response

    @action(detail=True, methods=['get', ],
            serializer_class=NoteSerializer,
            )
    def notes(self, request, pk=None):
        unit = Unit.objects.get(id=pk)
        self.context = {'request': request}
        serializer = NoteSerializer(unit.notes.all(), many=True, context=self.context)
        return Response(serializer.data)

    @transaction.atomic
    @action(detail=True, methods=['get', 'post'],
            serializer_class=UnitSerializer,
            )
    def link_files(self, request, pk=None):
        """
        This action will link an existing AzureFile to this unit.
        POST: {"id":1} to link azurfefile id=1 to this unit
        """
        self.context = {'request': request}
        queryset = Unit.objects.get(id=pk)
        if request.method == 'POST':
            params = json.loads(request.body)
            file_id = params.get('id')
            if file_id:
                attachment = AzureFile.objects.get(id=int(file_id))
                queryset.unit_images.add(attachment)
                queryset.save()
        serializer = UnitSerializer(queryset, many=False, context=self.context)
        return Response(serializer.data)

    # @transaction.atomic
    # @action(detail=True, methods=['get'],
    #     serializer_class=UnitSerializer,
    #     permission_classes=(ConfiguredPermission,)
    # )
    # def print_barcode(self, request, pk=None):
    #     import barcode
    #     from barcode.writer import ImageWriter
    #     image_buffer = BytesIO()
    #     image_buffer = barcode.get('ean13', '123456789102', writer=ImageWriter())
    #     # ean = barcode.get('ean13', '123456789102', writer=ImageWriter())
    #     image_buffer.save('foo')
    #     response = HttpResponse(image_buffer.getvalue(),content_type="image/png")
    #     return response
    def encode128(self, s):
        ''' Code 128 conversion for a font as described at
            https://en.wikipedia.org/wiki/Code_128 and downloaded
            from http://www.barcodelink.net/barcode-font.php
            Only encodes ASCII characters, does not take advantage of
            FNC4 for bytes with the upper bit set.
            It does not attempt to optimize the length of the string,
            Code B is the default to prefer lower case over control characters.
            Coded for https://stackoverflow.com/q/52710760/5987
        '''
        s = s.encode('ascii').decode('ascii')
        if s.isdigit() and len(s) % 2 == 0:
            # use Code 128C, pairs of digits
            codes = [105]
            for i in range(0, len(s), 2):
                codes.append(int(s[i:i + 2], 10))
        else:
            # use Code 128B and shift for Code 128A
            mapping = dict((chr(c), [98, c + 64] if c < 32 else [c - 32]) for c in range(128))
            codes = [104]
            for c in s:
                codes.extend(mapping[c])
        check_digit = (codes[0] + sum(i * x for i, x in enumerate(codes))) % 103
        codes.append(check_digit)
        codes.append(106)  # stop code
        chars = (b'\xd4' + bytes(range(33, 126 + 1)) + bytes(range(200, 211 + 1))).decode('latin-1')
        return ''.join(chars[x] for x in codes)

    @transaction.atomic
    @action(detail=True, methods=['get'],
            permission_classes=(AllowAny,)
            )
    def get_label(self, request, pk=None):
        from PIL import Image, ImageFont, ImageDraw
        from io import BytesIO
        from django.http import HttpResponse
        # make image of correct dpi/width/height
        # Customer project work orders
        # barcode Serial
        # Serial #
        # Wattage voltage
        self.context = {'request': request}
        encrypted = request.query_params.get('token')
        if encrypted:
            try:
                # this should be using Authenticate
                token = Token.objects.get(key=decrypt(encrypted))
                unit = Unit.objects.get(id=pk)
                # Don't know if I need any data... doing on POST for test
                # this printer is 8dpm Labels are 19x108
                dpi = 300
                width = 4 * dpi
                height = 1 * dpi
                sans_size = 45
                barcode_size = 70

                # barcode = ImageFont.truetype("./lsdb/fonts/LibreBarcode128-Regular.ttf", barcode_size)
                barcode = ImageFont.truetype("./lsdb/fonts/code128.ttf", barcode_size)
                mono = ImageFont.truetype("./lsdb/fonts/UbuntuMono-Regular.ttf", 50)
                sans = ImageFont.truetype("./lsdb/fonts/OpenSans-Bold.ttf", sans_size)
                serial = Image.new('1', (width, height), 1)
                draw = ImageDraw.Draw(serial)

                # title_text = '{} - {} - {}'.format(unit.project_set.first().customer.name,
                #                                    unit.project_set.first().number, unit.workorder_set.first().name)
                title_text = '{} - {} - {}'.format(unit.project_set.last().customer.name,
                                                   unit.project_set.last().number, unit.workorder_set.last().name)
                foo, mono_height = draw.textsize(unit.serial_number, font=mono)
                foo, title_height = draw.textsize(title_text, font=sans)

                # Make the barcode as big as will fit
                while True:
                    barcode_width, barcode_height = draw.textsize(unit.serial_number, font=barcode)
                    if (barcode_width >= width - 100) or (barcode_height >= height - 48 - mono_height - title_height):
                        break
                    else:
                        barcode_size += 4
                        # barcode = ImageFont.truetype("./lsdb/fonts/LibreBarcode128-Regular.ttf", barcode_size)
                        barcode = ImageFont.truetype("./lsdb/fonts/code128.ttf", barcode_size)
                encoded_serial = self.encode128(str(unit.serial_number))
                # print(encoded_serial)
                draw.text((width / 2, height / 2), encoded_serial, font=barcode, fill=0, anchor="mm")
                draw.text((width / 2, 20 + (height / 2) + (barcode_height / 2)), unit.serial_number, font=mono, fill=0,
                          anchor="mt")
                draw.text((width / 2, 12), title_text, font=sans, fill=0, anchor="mt")
                draw.text((12, height - 12),
                          '{:g}W'.format(unit.unit_type.module_property.nameplate_pmax),
                          font=sans, fill=0, anchor="ld")
                draw.text((width - 12, height - 12),
                          '{:g}V'.format(unit.unit_type.module_property.system_voltage),
                          font=sans, fill=0, anchor="rd")
                tsd = Image.new('1', (width, height), 1)
                draw = ImageDraw.Draw(tsd)
                # try:
                #     # if this blows up we're going to declare it a spare:
                #     tsd_text = unit.procedureresult_set.filter(test_sequence_definition__group__name__in=['Sequences','Control']
                #         ).first().test_sequence_definition.short_name
                # except:
                #     tsd_text = 'SPARE'
                # trap instead of using exception handling:
                if unit.procedureresult_set.filter(test_sequence_definition__group__name__in=['Sequences','Control']).count() < 1:
                    tsd_text = 'SPARE'
                else:
                    tsd_text = unit.procedureresult_set.filter(test_sequence_definition__group__name__in=['Sequences','Control']
                        ).first().test_sequence_definition.short_name
                # Make the tsd short_name as big as will fit
                while True:
                    tsd_width, tsd_height = draw.textsize(tsd_text, font=sans)
                    if (tsd_width >= width - 16) or (tsd_height >= height - 16):
                        break
                    else:
                        sans_size += 8
                        sans = ImageFont.truetype("./lsdb/fonts/OpenSans-Bold.ttf", sans_size)
                draw.text((width / 2, height / 2), tsd_text, font=sans, fill=0, anchor="mm")
                five_file = BytesIO()
                serial.save(five_file, "PDF", resolution=300.0, save_all=True,
                            append_images=[serial, serial, tsd, tsd])
                five_file.seek(0)
                filename = unit.serial_number  # self.default_filename
                self.content_type = 'application/pdf'
                response = HttpResponse(five_file, content_type=self.content_type)
                response['Content-Disposition'] = 'attachment; filename={}-labels.pdf'.format(filename)
                return response
            except:
                print('exception constructing PDF for label')
        return Response([])

    @action(detail=True, methods=['get'],
            # renderer_classes=(BinaryFileRenderer,),
            permission_classes=(ConfiguredPermission,),
            )
    def get_magic_link(self, request, pk=None):
        from rest_framework.renderers import JSONRenderer
        # The user has permission to get here, so I will send a link
        self.context = {'request': request}
        token = Token.objects.get(user=request.user)
        encrypted = encrypt(token.key)
        response = {'token': '{}'.format(encrypted)}
        json = JSONRenderer().render(response)
        return HttpResponse(json)

    @transaction.atomic
    @action(detail=True, methods=['get', 'post'],
            serializer_class=UnitSerializer,
            permission_classes=(ConfiguredPermission,)
            )
    def valid_asset(self, request, pk=None):
        """
        This action will test if the current unit has any reason to be at the provided asset.
        {
            "asset_name":"DiodeTestFixture_001",
            "procedure_definition":$ID <-- parameter for chamber entry to allow for test selection for qualification of the unit.
        }
        This will verify that there are procedure_result objects with disposition=null for the unit.
        If there are empty procedure_results, it will update the unit's fixture_location to reflect the unit has "checked in" at the  asset.
        procedure_results are returned in linear_execution_group order.
        It will return the asset object of there are NO results for this unit to fill in at this station's asset_type
        """
        self.context = {'request': request}
        unit = Unit.objects.get(id=pk)
        if request.method == 'POST':
            # Unit Disposition needs to be respected:
            if unit.disposition.complete:
                return Response({'message': 'This serial number not in an active disposition. Contact engineering.'},
                                status=HTTP_400_BAD_REQUEST)
            params = json.loads(request.body)
            asset = params.get('asset_name')
            definition = params.get('procedure_definition')
            if asset:
                # Has to account for calibrations return both
                # convert name string into asset object "BKL Wet Leakage Station",
                asset = Asset.objects.get(name__iexact=asset)
                if asset.disposition.name.upper() != 'AVAILABLE':
                    return Response({
                            'message': 'Asset {} is not available for use. Contact engineering.'.format(
                                asset)}, status=HTTP_400_BAD_REQUEST)
                # Check to see if this unit has *any* business at this asset type:
                if (asset.asset_types.filter(
                        id__in=AssetType.objects.filter(proceduredefinition__procedureresult__unit=unit))):
                    # Units "in_progress" can only check in at the current asset:
                    if unit.procedureresult_set.filter(work_in_progress_must_comply=True).exclude(
                            stepresult__measurementresult__asset__name=asset).count():
                        old_asset = unit.fixture_location.name
                        # EXCEPT: When there's more than one asset type?
                        return Response({
                            'message': 'This serial number marked IN PROGRESS at Asset {}. Contact engineering.'.format(
                                old_asset)}, status=HTTP_400_BAD_REQUEST)
                    unit.fixture_location = asset
                    unit.location = asset.location # This will have to change with deep location support
                    unit.save()
                    # Start building the queryset for sending back the procedure_result buckets
                    # One hit to DB here, not certain about the performnce hit of select_related()
                    # procedure_results = unit.procedureresult_set.filter(
                    #     Q(disposition__isnull=True) | Q(disposition__complete=False)).select_related(
                    #     'disposition','procedure_definition','test_sequence_definition','test_sequence_definition__group')
                    procedure_results = unit.procedureresult_set.filter(
                        Q(disposition__isnull=True) | Q(disposition__complete=False))
                    procedure_results = procedure_results.filter(
                        procedure_definition__asset_types__in=asset.asset_types.all())
                    procedure_results = procedure_results.exclude(supersede=True)
                    # procedure_results = procedure_results.order_by('linear_execution_group', '-allow_skip')
                    if definition:
                        procedure_results = procedure_results.filter(procedure_definition__id=definition)
                    # print('pre-ag count:',procedure_results.count())
                    # Not sure if either of these would benefit from select_related() or not
                    done_to = unit.procedureresult_set.aggregate(
                        done_to=Coalesce(
                            Max('linear_execution_group',
                                filter=Q(disposition__isnull=False,
                                test_sequence_definition__group__name__iexact='sequences',
                                supersede__isnull=True) | Q(disposition__isnull=False,
                                                         test_sequence_definition__group__name__iexact='sequences',
                                                         supersede=False)
                            ),0.0)).get('done_to')

                    if done_to != 0.0:
                        # this is NOT a virgin module, calibration or 0.0 EG case.
                        procedure_results = procedure_results.exclude(linear_execution_group__lt=done_to,
                            test_sequence_definition__group__name__iexact='sequences')

                    # this shouldn't hit the DB
                    highest_leg =unit.procedureresult_set.aggregate(
                            highest_leg = Coalesce(
                            Min('linear_execution_group',
                                filter=Q(disposition__isnull=True,
                                linear_execution_group__gt=done_to,
                                allow_skip=False,
                                test_sequence_definition__group__name__iexact='sequences',
                                supersede__isnull=True) | Q(disposition__isnull=True,
                                                            allow_skip=False,
                                                            linear_execution_group__gt=done_to,
                                                            test_sequence_definition__group__name__iexact='sequences',
                                                            supersede=False)
                            ), 99.0)).get('highest_leg')
                    # print('',done_to, highest_leg)
                    procedure_results = procedure_results.exclude(linear_execution_group__gt=highest_leg,
                        test_sequence_definition__group__name__iexact='sequences')
                    # MD: One more pass at this query might be needed, I think I can do away with the annotations
                    # using intersection() but I want to get this code running.
                    serializer = ProcedureResultSerializer(procedure_results, many=True, context=self.context)
                    # MD: Add testing directives to an object in serializer.data here:
                    return Response(serializer.data)
                # TODO: Need a better return behavior here: return empty set for now
                return Response([])
        # GET supported only for DRF views
        else:
            serializer = UnitSerializer(unit, many=False, context=self.context)
        return Response(serializer.data)
