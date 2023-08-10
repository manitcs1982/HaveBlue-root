import json
import fileinput
import os
from openpyxl import load_workbook
from datetime import datetime

from django.db import IntegrityError, transaction
from django.core.exceptions import ObjectDoesNotExist
from django.apps import apps

from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from datetime import datetime, timedelta, date

from lsdb.models import AvailableDefect
from lsdb.models import Crate
from lsdb.models import Customer
from lsdb.models import Disposition
from lsdb.models import Group
from lsdb.models import Location
from lsdb.models import MeasurementDefinition
from lsdb.models import ModuleProperty
from lsdb.models import ModuleTechnology
from lsdb.models import Note
from lsdb.models import PlexusImport
from lsdb.models import ProcedureDefinition
from lsdb.models import ProcedureResult
from lsdb.models import Project
from lsdb.models import Unit
from lsdb.models import StepDefinition
from lsdb.models import StepResult
from lsdb.models import MeasurementResult
from lsdb.models import TestSequenceDefinition
from lsdb.models import UnitType
from lsdb.models import UnitTypeFamily
from lsdb.models import UserProfile
from lsdb.models import UserRegistrationStatus
from lsdb.models import WorkOrder

from lsdb.views import WorkOrderViewSet


class Command(BaseCommand):
    help = 'Loads serial numbers from Ryan\'s excel file into  lsdb'
    me = User.objects.get(pk=1)
    workbook = load_workbook("./lsdb/fixtures/loader/Serial Number to work order-test sequence.xlsx")
    # workbook = load_workbook("./lsdb/fixtures/loader/Boviet.xlsx")
    serials = workbook.active
    lsdb = apps.get_app_config('lsdb')
    auth = apps.get_app_config('auth')

    def handle(self, *args, **options):
        for row in self.serials.iter_rows():
            row_data={}
            for count, cell in enumerate(row):
                row_data[count]=cell.value
            if row_data[0] == "UnitID":
                continue
            if row_data[4] == "#N/A":
                continue

            try:
                unit = Unit.objects.get(pk=row_data[0])
                self.unit = unit
            except ObjectDoesNotExist:
                print('WTF')
                raise ObjectDoesNotExist
            try:
                customer = Customer.objects.get(name=row_data[2])
            except ObjectDoesNotExist:
                print('customer {} Not found'.format(row_data[2]))
                continue
            # need to find the work order that this unit is attached to:
            if unit.project_set.count() == 1:
                project = unit.project_set.all()[0]
                # print('{}'.format(project.workorder_set.all()))
            else:
                continue
                print('{}'.format(unit.project_set.all()))
            if unit.procedureresult_set.count()>0:
                print('{} found {} results, skipping'.format(unit.serial_number, unit.procedureresult_set.count()))
                continue

            try:
                work_order = project.workorder_set.get(name=row_data[4])
            except ObjectDoesNotExist:
                # print('unit {} in project {} does not have workorder {} Trying alt'.format(
                #     unit.serial_number,
                #     project.number,
                #     row_data[4]
                # ))
                try:
                    work_order = project.workorder_set.get(name__iexact='BOM {}'.format(row_data[4]))
                except ObjectDoesNotExist:
                    print('unit {} in {} project {} does not have workorder {} giving up'.format(
                        unit.serial_number,
                        customer.name,
                        project.number,
                        'BOM {}'.format(row_data[4])
                    ))
                    continue
            # Next we need to find the test sequence:
            test_sequence = work_order.test_sequence_definitions.filter(name__icontains=row_data[5])
            if test_sequence.count() == 0:
                    print('unit {} in {} project {} in workorder {} want: {} have:{}'.format(
                        unit.serial_number,
                        customer.name,
                        project.number,
                        'BOM {}'.format(row_data[4]),
                        row_data[5],
                        work_order.test_sequence_definitions.all()
                    ))
                    #should we try the other project?
            elif test_sequence.count() == 1:
                print('{} has a match!'.format(unit.serial_number))
                # Don't do this more than once:
                self.test_sequence = test_sequence[0]

    @transaction.atomic
    def do_write(self):
        test_sequence = self.test_sequence
        unit = self.unit
        for execution in test_sequence.procedureexecutionorder_set.all():
            # print(procedure.procedureexecutionorder_set.all()[0])
            procedure_result = ProcedureResult.objects.create(
                unit = unit,
                name = execution.execution_group_name,
                disposition = None,
                group = execution.procedure_definition.group,
                work_order = work_order,
                procedure_definition = execution.procedure_definition,
                version = execution.procedure_definition.version,
                linear_execution_group = execution.execution_group_number,
                test_sequence_definition = test_sequence,
                allow_skip = True,
            )
            for step_execution in execution.procedure_definition.stepexecutionorder_set.all():
                step_result = StepResult.objects.create(
                    name = step_execution.step_definition.name,
                    procedure_result = procedure_result,
                    step_definition = step_execution.step_definition,
                    execution_number = 0,
                    disposition = None,
                    start_datetime = None,
                    duration = 0,
                    test_step_result = None,
                    archived = False,
                    description = None,
                    step_number = 0,
                    step_type = step_execution.step_definition.step_type,
                    linear_execution_group = step_execution.execution_group_number,
                    allow_skip = True,
                )
                for measurement_definition in step_execution.step_definition.measurementdefinition_set.all():
                    measurement_result = MeasurementResult.objects.create(
                        step_result = step_result,
                        measurement_definition = measurement_definition,
                        software_revision = 0.0,
                        disposition = None,
                        limit = measurement_definition.limit,
                        station = 0,
                        name = measurement_definition.name,
                        record_only = measurement_definition.record_only,
                        allow_skip = measurement_definition.allow_skip,
                        requires_review = measurement_definition.requires_review,
                        measurement_type = measurement_definition.measurement_type,
                        order = measurement_definition.order,
                        report_order = measurement_definition.report_order,
                        measurement_result_type = measurement_definition.measurement_result_type,
                    )

    def load_modules(self):
        skipped =0
        self.counts['PVModule'] =0
        with fileinput.input(files=('./lsdb/fixtures/loader/PVModule.json')) as f:
            for line in f:
                obj = json.loads(line)
                id = obj.get('_id')
                try:
                    plexus_oid = self.get_oid(plexus_oid=id['$oid'])
                    skipped +=1
                    print('skipping')
                    continue
                except ObjectDoesNotExist:
                    pass
                try:
                    if obj.get('originalCrate'):
                        crate = self.get_oid(plexus_oid=obj.get('originalCrate').get("$oid"))
                    else:
                        crate = self.crate
                    object = Unit.objects.create(
                        unit_type = self.get_oid(plexus_oid=obj.get('moduleType').get("$oid")),
                        crate = crate,
                        serial_number = obj.get('serialNumber'),
                        location = self.location,
                    )

                    if obj.get('quote'):
                        # Cowardly not creating Work Orders without a project
                        if obj.get('BOM'):
                            try:
                                work_order = WorkOrder.objects.get(
                                    name=obj.get('BOM'),
                                    project=self.get_oid(plexus_oid=obj.get('quote').get('$oid'))
                                )
                            except ObjectDoesNotExist:
                                work_order = WorkOrder.objects.create(
                                    name = obj.get('BOM'),
                                    project = self.get_oid(plexus_oid=obj.get('quote').get('$oid')),
                                    disposition = self.placeholder,
                                )
                            work_order.units.add(object)
                        self.get_oid(plexus_oid=obj.get('quote').get('$oid')).units.add(object)
                        plexus_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                            lsdb_model = 'unit',
                            lsdb_id = object.id
                        )
                    self.counts['PVModule'] += 1
                except Exception as e:
                    skipped +=1
                    print(e)
                    print(obj)
        print('{} modules loaded, {} skipped'.format(self.counts['PVModule'],skipped))


    def load_projects(self):
        skipped = 0
        self.counts['Quote'] = 0
        with fileinput.input(files=('./lsdb/fixtures/loader/Quote.json')) as f:
            for line in f:
                obj = json.loads(line)
                id = obj.get('_id')
                try:
                    plexus_oid = self.get_oid(plexus_oid=id['$oid'])
                    skipped +=1
                except ObjectDoesNotExist:
                    try:
                        object = Project.objects.create(
                            number = obj.get('name'),
                            customer = self.get_oid(plexus_oid= obj.get("customer")["$oid"]),
                            project_manager = self.me,
                            group=self.project_group,
                            disposition = self.placeholder,
                            # start_date = obj.get('created_on')['$date'],
                        )
                        plexus_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                            lsdb_model = 'project',
                            lsdb_id = object.id
                        )
                        self.counts['Quote'] +=1
                    except IntegrityError:
                        # need to deduplicate these too...
                        object = Project.objects.get(
                            number = obj.get('name'),
                            group=self.project_group,
                        )
                        plexus_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                            lsdb_model = 'project',
                            lsdb_id = object.id
                        )

        print('{} projects loaded, {} skipped'.format(self.counts['Quote'],skipped))

    def load_crates(self):
        skipped =0
        self.counts['PVModuleCrate'] = 0
        with fileinput.input(files=('./lsdb/fixtures/loader/PVModuleCrate.json')) as f:
            for line in f:
                obj = json.loads(line)
                id = obj.get('_id')
                try:
                    plexus_oid = self.get_oid(plexus_oid=id['$oid'])
                    skipped +=1
                except ObjectDoesNotExist:
                    try:
                        object = Crate.objects.create(
                            name = obj.get('name'),
                            disposition = self.placeholder,
                            shipped_by = self.get_oid(plexus_oid=obj.get("shippedBy")["$oid"])
                        )
                        crate_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                            lsdb_model = 'crate',
                            lsdb_id = object.id
                        )
                    except IntegrityError:
                        # deduplicate crate names
                        object = Crate.objects.get(
                            name = obj.get('name'),
                            shipped_by = self.get_oid(plexus_oid=obj.get("shippedBy")["$oid"])
                        )
                        crate_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                            lsdb_model = 'crate',
                            lsdb_id = object.id
                        )
                    try:
                        self.counts['PVModuleCrate'] +=1
                        # Notes children can be created now, photos, need special handling
                        for observation in obj.get('observations'):
                            if observation.get('photo'):
                                # TODO: This will be handled in a separate pass
                                # print(observation)
                                pass
                            if observation.get('comment'):
                                note_id = observation.get('_id')
                                note = Note.objects.create(
                                    user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                                    subject = "Observation Comment",
                                    text = observation.get('comment'),
                                    datetime = obj.get('created on')['$date'],
                                )
                                object.observations.add(note)
                                object.save()
                                note_oid = self.store_oid(
                                    plexus_oid = note_id['$oid'],
                                    user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                                    lsdb_model = 'note',
                                    lsdb_id = note.id
                                )
                    except IntegrityError as e:
                        print(e)
                        self.oids[obj.get('_id')['$oid']] = Crate.objects.get(name = obj.get('name'),
                                shipped_by = self.oids[obj.get("shippedBy")["$oid"]])
                        # skipped +=1
                    # print(e)
        print('{} crates loaded, {} skipped'.format(self.counts['PVModuleCrate'],skipped))

    def load_types(self):
        skipped =0
        self.counts['PVModuleType'] = 0
        with fileinput.input(files=('./lsdb/fixtures/loader/PVModuleType.json')) as f:
            for line in f:
                obj = json.loads(line)
                id = obj.get('_id')
                try:
                    plexus_oid = self.get_oid(plexus_oid=id['$oid'])
                    skipped +=1
                    continue
                except ObjectDoesNotExist:
                    try:
                        # get a unique moduleType to see if it exists before blindly making another one:
                        # print(obj.get('manufacturer').get('$oid'))
                        # mfr_oid = obj.get('manufacturer').get('$oid')
                        object = UnitType.objects.get(
                            model = obj.get('modelNumber'),
                            manufacturer = self.get_oid(plexus_oid=obj.get('manufacturer').get('$oid'))
                        )
                        # put the oid in:
                        plexus_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                            lsdb_model = 'unittype',
                            lsdb_id = object.id
                        )
                        continue
                    except ObjectDoesNotExist:
                        pass #?
                try:
                    mod_props = ModuleProperty.objects.create(
                        number_of_cells = obj.get('cellData',{'cellsInSeries':0}).get('cellsInSeries',0),
                        nameplate_pmax = obj.get('nameplateData',{'pmax':0}).get('pmax',0),
                        module_width = obj.get('moduleDimensions',{'width':0}).get('width',0),
                        module_height = obj.get('moduleDimensions',{'height':0}).get('height',0),
                        system_voltage = obj.get('systemVoltage',0),
                        module_technology = self.default_tech,
                        isc = obj.get('nameplateData',{'isc':0}).get('isc',0),
                        voc = obj.get('nameplateData',{'voc':0}).get('voc',0),
                        imp = obj.get('nameplateData',{'imp':0}).get('imp',0),
                        vmp = obj.get('nameplateData',{'vmp':0}).get('vmp',0),
                    )
                    if obj.get('cellData'):
                        module_technology = ModuleTechnology.objects.get(name = obj.get('cellData').get('cellType',"Mono-PERC"))
                        mod_props.cells_in_series = obj.get('cellData').get('cellsInSeries',0)
                        mod_props.cells_in_parallel = obj.get('cellData').get('cellsInParallel',0)
                        mod_props.cell_area = obj.get('cellData').get('cellArea',0)
                        mod_props.save()
                    if obj.get('temperatureCoefficients'):
                        mod_props.alpha_isc = obj.get('temperatureCoefficients').get('alphaIsc',0)
                        mod_props.beta_voc = obj.get('temperatureCoefficients').get('betaVoc',0)
                        mod_props.amma_pmp = obj.get('temperatureCoefficients').get('gammaPmax',0)
                        mod_props.save()
                    object = UnitType.objects.create(
                        model = obj.get('modelNumber'),
                        manufacturer = self.get_oid(plexus_oid=obj.get('manufacturer').get('$oid')),
                        unit_type_family = self.unit_type_family,
                        module_property = mod_props,
                    )
                    plexus_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.get_oid(plexus_oid=obj.get("created by")["$oid"]),
                            lsdb_model = 'unittype',
                            lsdb_id = object.id
                    )
                    self.counts['PVModuleType'] +=1
                except Exception as e:
                    print(e)
                    skipped +=1
        print('{} Types loaded, {} skipped'.format(self.counts['PVModuleType'],skipped))

    def load_users(self):
        count = 0
        skipped = 0
        status = UserRegistrationStatus.objects.get(status="New User")
        with fileinput.input(files=('./lsdb/fixtures/loader/plexus_users.json')) as f:
            for line in f:
                obj = json.loads(line)
                id = obj.get('_id')
                try:
                    plexus_oid = self.get_oid(plexus_oid=id['$oid'])
                    skipped +=1
                except ObjectDoesNotExist:
                    user = User.objects.create(
                        username = obj.get('email'),
                        is_active=True,
                        first_name =obj.get('first name'),
                        last_name = obj.get('last name'),
                    )
                    user.set_password('solar123')
                    user.save()
                    UserProfile.objects.create(
                        user=user,
                        notes = id['$oid'],
                        user_registration_status=status,
                        )
                    plexus_oid = self.store_oid(
                        plexus_oid = id['$oid'],
                        user = self.me,
                        lsdb_model = 'user',
                        lsdb_id = user.id
                    )
                    count +=1
        print('{} users loaded, {} skipped'.format(count,skipped))

    def load_defects(self):
        count = 0
        skipped = 0
        self.counts['PVVisualInspectionDefect'] = 0
        with fileinput.input(files=('./lsdb/fixtures/loader/PVVisualInspectionDefect.json')) as f:
            for line in f:
                obj = json.loads(line)
                id = obj.get('_id')
                try:
                    plexus_oid = self.get_oid(plexus_oid=id['$oid'])
                    skipped +=1
                except ObjectDoesNotExist:
                    try:
                        object = AvailableDefect.objects.create(
                            category = obj.get('category'),
                            description = obj.get('description'),
                            short_name = obj.get('shortName'),
                        )
                    except IntegrityError:
                        try:
                            defect = AvailableDefect.objects.get(category=obj.get('category'), short_name=obj.get('shortName'))
                            plexus_oid = self.store_oid(
                                plexus_oid = id['$oid'],
                                user = self.me,
                                lsdb_model = 'availabledefect',
                                lsdb_id = defect.id
                            )
                        except IntegrityError:
                            pass
                    count +=1
        print('{} defects loaded, {} skipped'.format(count,skipped))


    def load_customers(self):
        skipped =0
        self.counts['Corporation'] = 0
        with fileinput.input(files=('./lsdb/fixtures/loader/Corporation.json')) as f:
            for line in f:
                obj = json.loads(line)
                type = obj.get('__t')
                id = obj.get('_id')
                try:
                    plexus_oid = self.get_oid(plexus_oid=id['$oid'])
                    skipped +=1
                except ObjectDoesNotExist:
                    try:
                        object = Customer.objects.create(
                            name = obj.get('name')
                        )
                    except IntegrityError:
                        # this deduplicates while maintaining all references to the single
                        object = Customer.objects.get(
                            name = obj.get('name')
                        )
                        plexus_oid = self.store_oid(
                            plexus_oid = id['$oid'],
                            user = self.me,
                            lsdb_model = 'customer',
                            lsdb_id = object.id
                        )
                    self.counts['Corporation'] +=1
        print('{} customers loaded, {} skipped'.format(self.counts['Corporation'],skipped))

    def get_counts(self): # stubbed
        # dump record counts for the lulz
        with fileinput.input(files=('./lsdb/fixtures/loader/plexus_usermgds.json')) as f:
            for line in f:
                obj = json.loads(line)
                type = obj.get('__t')

                type_count = self.types.get(type)
                if type_count:
                    self.types[type] += 1
                else:
                    self.types[type] = 1

        for type in self.types.keys():
            print('{} : {} / {}'.format(type, self.types[type],
                self.counts.get(type,0)
                ))
